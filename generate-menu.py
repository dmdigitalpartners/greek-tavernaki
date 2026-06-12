"""
Generate menu-data.js from Restaurant_Menu.xlsx + Master_Menu_Database.xlsx.
Run once: python3 generate-menu.py
Commit the output (menu-data.js) to the repo — do NOT re-run in production.
"""

import re
import openpyxl

# ---------------------------------------------------------------------------
# Transliteration: Bulgarian Cyrillic → ASCII slug
# ---------------------------------------------------------------------------
_BG_MAP = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l',
    'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's',
    'т': 't', 'у': 'u', 'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch',
    'ш': 'sh', 'щ': 'sht', 'ъ': 'a', 'ь': '', 'ю': 'yu', 'я': 'ya',
    'ы': 'i',
}

def _transliterate(text):
    out = []
    for ch in text.lower():
        if ch in _BG_MAP:
            out.append(_BG_MAP[ch])
        else:
            out.append(ch)
    return ''.join(out)

def slugify(text):
    text = _transliterate(str(text))
    text = re.sub(r'[^a-z0-9]+', '-', text)
    text = text.strip('-')
    text = re.sub(r'-+', '-', text)
    return text

# ---------------------------------------------------------------------------
# Category slug mapping (normalises typos in the xlsx)
# ---------------------------------------------------------------------------
CATEGORY_MAP = {
    'Салати / Salads':                              'salads',
    'Тартери / Starters':                           'starters',
    'Предястия / Appetizers':                       'appetizers',
    'Мезета / Starters Plate':                      'mezethes',
    'Пържени Картофи / Fried Potatoes':             'potatoes',
    'Паста и Ризото / Pasta and Risotto':           'pasta',
    'Паста и Ризото / Pasta и Risotto':             'pasta',   # typo variant
    'Месни Ястия на Скара / BBQ Meat Dishes':       'bbq-meat',
    'Месни Ястия на Тиган / Pan Fried Meat Dishes': 'pan-meat',
    'Специалитети / Specialties':                   'specialties',
    'Риба и Морски Дарове / Fish and Seafood':      'fish',
    'Риба и Morton Дарове / Fish and Seafood':      'fish',    # OCR typo variant
}

# ---------------------------------------------------------------------------
# Per-weight items (price is per 100g, not fixed)
# Manually verified from xlsx — large whole fish sold by weight
# ---------------------------------------------------------------------------
PER_WEIGHT_IDS = {
    'bbq-lavrak-2kg',        # BBQ Лаврак 2кг   — EUR 7.62 / 100g
    'bbq-red-sea-bass',      # BBQ Red Sea Bass  — EUR 6.08 / 100g
    'dzha-wild-bbq-lavrak-fari-tsipura',  # ДЖА Wild BBQ — EUR 7.62 / 100g
}

# ---------------------------------------------------------------------------
# Image assignments (brand_assets/ filenames → BG dish names they map to)
# ---------------------------------------------------------------------------
IMAGE_BY_NAME_BG = {
    'Киноа и Скариди':              'brand_assets/kinoa-skaridi.png',
    'Плато Морски Дарове':          'brand_assets/plato-morski-darove.png',
    'Свински Бон Филе с Лютти Чушки': 'brand_assets/svinski-kascheta-chushki.png',
}

# ---------------------------------------------------------------------------
# Lunch items: English translations (Master_Menu_Database has no EN column)
# ---------------------------------------------------------------------------
LUNCH_EN = {
    'Гръцка мусака':                          'Greek Moussaka',
    'Зеле със свинско':                       'Cabbage with Pork',
    'Бъркани яйца':                           'Scrambled Eggs with Cheese & Tomatoes',
    'Пържени тиквички с дзадзики (обяд)':    'Fried Zucchini with Tzatziki (Lunch)',
    'Таратор':                                'Tarator (Cold Yoghurt Soup)',
    'Пилешка супа':                           'Chicken Soup',
    'Салата с катък, домати и краставици':    'Salad with Katyk, Tomatoes & Cucumbers',
    'Салата Рока марули':                     'Rocket & Lettuce Salad',
    'Салата зеле и моркови (обяд)':           'Cabbage & Carrots Salad (Lunch)',
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_eur(val):
    """Return float EUR price from various cell formats."""
    if val is None:
        return None
    s = str(val).replace('€', '').replace(' ', '').strip()
    try:
        return round(float(s), 2)
    except ValueError:
        return None

def format_weight(val):
    """Return '400g' string or None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == 'None':
        return None
    # Strip existing 'g' suffix if already present
    s = s.rstrip('g').strip()
    try:
        g = int(float(s))
        return f'{g}g'
    except ValueError:
        return s + 'g' if s else None

def make_id(name_bg, seen_ids):
    """Unique slug from Bulgarian name."""
    base = slugify(name_bg)
    candidate = base
    n = 2
    while candidate in seen_ids:
        candidate = f'{base}-{n}'
        n += 1
    seen_ids.add(candidate)
    return candidate

# ---------------------------------------------------------------------------
# Read Restaurant_Menu.xlsx — 106 items
# Columns (0-indexed): Category, Name_BG, Name_EN, Price_BGN, Price_EUR,
#                      Weight_g, Ingredients, Description_EN, Allergens, Notes
# ---------------------------------------------------------------------------
def read_restaurant_menu():
    wb = openpyxl.load_workbook('Restaurant_Menu.xlsx')
    ws = wb['Menu']
    rows = []
    seen_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat_raw, name_bg, name_en, _, price_eur_raw, weight_raw, _, _, allergens_raw, _ = row
        if not cat_raw or not name_bg:
            continue
        cat_raw = str(cat_raw).strip()
        category = CATEGORY_MAP.get(cat_raw)
        if category is None:
            print(f'  WARNING: unknown category "{cat_raw}" — skipped')
            continue
        name_bg = str(name_bg).strip()
        name_en = str(name_en).strip() if name_en else name_bg
        price = parse_eur(price_eur_raw)
        weight = format_weight(weight_raw)
        allergens = str(allergens_raw).strip() if allergens_raw else 'Requires Restaurant Verification'
        dish_id = make_id(name_bg, seen_ids)
        price_unit = '100g' if dish_id in PER_WEIGHT_IDS else None
        image = IMAGE_BY_NAME_BG.get(name_bg)
        rows.append({
            'id':         dish_id,
            'category':   category,
            'name_bg':    name_bg,
            'name_en':    name_en,
            'price':      price,
            'price_unit': price_unit,
            'weight':     weight,
            'allergens':  allergens,
            'image':      image,
        })
    return rows, seen_ids

# ---------------------------------------------------------------------------
# Read Master_Menu_Database.xlsx — 9 lunch items
# Columns (0-indexed): Category, Product_Name, Price_EUR, Grams, Description, ...
# ---------------------------------------------------------------------------
def read_lunch_menu(seen_ids):
    wb = openpyxl.load_workbook('Master_Menu_Database.xlsx')
    ws = wb['Master Menu']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cat_raw, name_bg, price_raw, weight_raw = row[0], row[1], row[2], row[3]
        if not cat_raw or str(cat_raw).strip() != 'Обедно меню':
            continue
        if not name_bg:
            continue
        name_bg = str(name_bg).strip()
        name_en = LUNCH_EN.get(name_bg, name_bg)
        price = parse_eur(price_raw)
        weight = format_weight(weight_raw)
        allergens = 'Requires Restaurant Verification'
        dish_id = make_id(name_bg, seen_ids)
        rows.append({
            'id':         dish_id,
            'category':   'lunch',
            'name_bg':    name_bg,
            'name_en':    name_en,
            'price':      price,
            'price_unit': None,
            'weight':     weight,
            'allergens':  allergens,
            'image':      None,
        })
    return rows

# ---------------------------------------------------------------------------
# Serialise to JS
# ---------------------------------------------------------------------------
def js_val(v):
    if v is None:
        return 'null'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return str(v)
    escaped = str(v).replace('\\', '\\\\').replace("'", "\\'")
    return f"'{escaped}'"

def write_menu_data_js(items):
    lines = [
        '// AUTO-GENERATED by generate-menu.py — do not edit manually',
        'window.menuData = [',
    ]
    for item in items:
        price_str = f'{item["price"]:.2f}' if item['price'] is not None else 'null'
        lines.append('  {')
        lines.append(f'    id:         {js_val(item["id"])},')
        lines.append(f'    category:   {js_val(item["category"])},')
        lines.append(f'    name_bg:    {js_val(item["name_bg"])},')
        lines.append(f'    name_en:    {js_val(item["name_en"])},')
        lines.append(f'    price:      {price_str},')
        lines.append(f'    price_unit: {js_val(item["price_unit"])},')
        lines.append(f'    weight:     {js_val(item["weight"])},')
        lines.append(f'    allergens:  {js_val(item["allergens"])},')
        lines.append(f'    image:      {js_val(item["image"])},')
        lines.append('  },')
    lines.append('];')
    with open('menu-data.js', 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print('Reading Restaurant_Menu.xlsx...')
    restaurant_items, seen_ids = read_restaurant_menu()
    print(f'  {len(restaurant_items)} items loaded')

    print('Reading Master_Menu_Database.xlsx (lunch)...')
    lunch_items = read_lunch_menu(seen_ids)
    print(f'  {len(lunch_items)} lunch items loaded')

    all_items = restaurant_items + lunch_items

    # Summary
    from collections import Counter
    by_cat = Counter(d['category'] for d in all_items)
    per_weight = [d['id'] for d in all_items if d['price_unit'] == '100g']
    images = [d['image'] for d in all_items if d['image']]

    print('\n=== SUMMARY ===')
    print(f'Total items: {len(all_items)}')
    print('Per category:')
    expected = {
        'salads': 12, 'starters': 3, 'appetizers': 13, 'mezethes': 8,
        'potatoes': 6, 'pasta': 11, 'bbq-meat': 13, 'pan-meat': 9,
        'specialties': 8, 'fish': 23, 'lunch': 9,
    }
    for cat, exp in expected.items():
        actual = by_cat.get(cat, 0)
        status = '✓' if actual == exp else f'✗ EXPECTED {exp}'
        print(f'  {cat}: {actual} {status}')
    print(f'Per-weight items ({len(per_weight)}): {per_weight}')
    print(f'Image paths assigned ({len(images)}): {images}')

    # Check per-weight IDs actually matched
    matched_pw = set(per_weight)
    unmatched = PER_WEIGHT_IDS - matched_pw
    if unmatched:
        print(f'\n  WARNING: PER_WEIGHT_IDS not matched in data: {unmatched}')
        print('  Actual fish IDs:')
        for d in all_items:
            if d['category'] == 'fish':
                print(f'    {d["id"]}  ← {d["name_bg"]}')

    print('\nWriting menu-data.js...')
    write_menu_data_js(all_items)
    print('Done.')

if __name__ == '__main__':
    main()
